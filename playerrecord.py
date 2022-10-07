import os
import sys
import pandas as pd
import mymodule
import subprocess
import openpyxl as px
from openpyxl.styles import Alignment, PatternFill
from itertools import accumulate

def get_game_metadata(file_path):
    _game = file_path.split('\\')[-1]
    game  = _game.split('_')
    return game[0], game[1], game[2].split('.')[0]

def insert_game_metadata(df, date, style, team):
    df.insert(0,'日付',date)
    df.insert(1,'試合形式',style)
    df.insert(2,'チーム',team)

# 試合結果を統合
def concat_games(paths, df_bat, df_pitch):
    for path in paths:
        date, style, team = get_game_metadata(path)
        # 打撃
        df_bat_read_excel=pd.read_excel(path, sheet_name='打撃成績')
        insert_game_metadata(df_bat_read_excel, date, style, team)
        df_bat=pd.concat([df_bat, df_bat_read_excel])
        # 投手
        df_pitch_read_excel=pd.read_excel(path, sheet_name='投手成績')
        insert_game_metadata(df_pitch_read_excel, date, style, team)
        df_pitch=pd.concat([df_pitch, df_pitch_read_excel])
    return df_bat, df_pitch

# 列の幅を設定
def set_column_width(ws: px.Workbook.worksheets, beginning_rate: int):
    date_width = 12
    style_width = 12
    team_width = 20
    width = 5
    
    ws.column_dimensions['A'].width = date_width
    ws.column_dimensions['B'].width = style_width
    ws.column_dimensions['C'].width = team_width
    
    for col in ws.iter_cols(min_col=4):
        column_num = col[0].column
        column_str = col[0].column_letter
        if column_num == beginning_rate:
            width = 7
        ws.column_dimensions[column_str].width = width

# 一行目の書式を設定
def set_format_row1(ws: px.Workbook.worksheets):
    for row in ws.iter_rows(min_col=4,max_row=1):
        for cell in row:
            cell.alignment = Alignment(vertical='center', textRotation=255)
    ws['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
    ws['B1'].alignment = Alignment(horizontal = 'center', vertical='center')
    ws['C1'].alignment = Alignment(horizontal = 'center', vertical='center')      

# インデックス（試合形式，チーム）の書式を設定
def set_format_index(ws: px.Workbook.worksheets):
    for col in ws.iter_cols(min_col=2, max_col=3, min_row=2, max_row=ws.max_row-1):
        for cell in col:
            cell.alignment = Alignment(horizontal = 'left')

#　セルの背景色を設定
def set_backgroud_color(ws: px.Workbook.worksheets, title_color: str, name_color: str):
    max_row = ws.max_row
    # title
    for row in ws.iter_rows(max_row=1):
        for cell in row:
            cell.fill = PatternFill(patternType = 'solid', fgColor=title_color)
    for row in ws.iter_rows(min_col = 2,min_row=max_row):
        for cell in row:
            cell.fill = PatternFill(patternType = 'solid', fgColor=name_color)        
    
    
# 書式設定　まとめ
def set_format(ws: px.Workbook.worksheets, games, beginning_rate, deep_color, thin_color):
    ws['B'+str(ws.max_row)]=str(games)+'試合'
    ws['C'+str(ws.max_row)]='合計'
        
    # 列と行を固定化
    ws.freeze_panes = 'D2'
    set_column_width(ws, beginning_rate)
    mymodule.set_rate_format(ws, beginning_rate)
    set_format_row1(ws)
    set_format_index(ws)
    set_backgroud_color(ws, deep_color, thin_color)    

# グラフ作成
def make_avg_graph_sheet(ws_bat: pd.DataFrame, ws_avg: pd.DataFrame):
    
    # カラムを設定
    columns = ['日付', '打数', '安打', '打率', '通算']
    i = 1
    for column in columns:
        ws_avg.cell(row=1, column=i).value = column
        i+=1
    
    #とりあえず日付をコピー（セル結合は無視）
    i = 2
    for col in ws_bat.iter_cols(min_row=2, max_col=1):
        for cell in col:
                ws_avg.cell(row=i,column=1).value = cell.value
                i+=1

    #NANの部分を補填
    i = 2
    for col in ws_avg.iter_cols(min_row=2,max_col=1,max_row=ws_bat.max_row-1):
        for cell in col:
            if cell.value is None:
                ws_avg.cell(row=i,column=1).value = ws_avg.cell(row=i-1,column=1).value
            i+=1
    
    at_bats = []
    hits = []
    at_bats_col_num = 5
    hits_col_num = 6
    at_bats_cumulative = []
    hits_cumulative = []
    
    for col in ws_bat.iter_cols(min_row=2, min_col=5, max_row=ws_bat.max_row-1, max_col=6):
        for cell in col:
            if cell.column == at_bats_col_num:
                at_bats.append(cell.value)
            elif cell.column == hits_col_num:
                hits.append(cell.value)
    at_bats_cumulative = list(accumulate(at_bats))
    hits_cumulative = list(accumulate(hits))
    
    # 打率推移シートに累積和を渡す
    for i in range(ws_bat.max_row-2):
        ws_avg.cell(row=i+2, column=2).value = at_bats_cumulative[i]
        ws_avg.cell(row=i+2, column=3).value = hits_cumulative[i]
        
    for i in range(ws_bat.max_row-2):
        if at_bats_cumulative[i] == 0:
            avg = 0.000
        else:
            avg = hits_cumulative[i]/at_bats_cumulative[i]
        ws_avg.cell(row=i+2, column=4).value = avg
      
    if at_bats_cumulative[-1] == 0:
        total_avg = 0.000
    else:
        total_avg = hits_cumulative[-1]/at_bats_cumulative[-1]
    for i in range(ws_bat.max_row-2):
        ws_avg.cell(row=i+2, column=5).value = total_avg

    chart = px.chart.LineChart()
    data = px.chart.Reference(ws_avg, min_col=4, max_col=5, min_row=1, max_row=ws_avg.max_row-1)
    categories = px.chart.Reference(ws_avg, min_col=1, max_col=1, min_row=2, max_row=ws_avg.max_row-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.style = 34
    chart.title='打率の推移'
    chart.height=9
    chart.width=16
    chart.x_axis.title = '日付'

    ws_avg.add_chart(chart, "G2")


def main():
    
    BAT_DARK_COLOR = 'A4C6FF'
    BAT_THIN_COLOR = 'D9E5FF'

    PITCH_DARK_COLOR = 'FFA3A3'
    PITCH_THIN_COLOR = 'FFD9D9'
    
    beginning_bat_rate = 26
    beginning_pitch_rate = 19
    
    print("start")
    
    # パスを設定
    dirname = mymodule.get_dirname()
    import_foloder_path = dirname + '\試合結果'
    export_folder_path  = dirname + '\個人成績'
    
     # 試合結果の全データを統合
    df_bat_concat   = pd.DataFrame()
    df_pitch_concat = pd.DataFrame()
    game_file_paths = mymodule.get_xlsx_file_paths(folder_path=import_foloder_path)
    df_bat_concat, df_pitch_concat = concat_games(game_file_paths, df_bat_concat, df_pitch_concat)
    
    # 成績を生成する選手を抽出
    players_name = mymodule.get_players_name(dirname)
    df_bat_concat   = df_bat_concat[df_bat_concat['名前'].isin(players_name)]
    df_pitch_concat = df_pitch_concat[df_pitch_concat['名前'].isin(players_name)]
    
    #個人の名前を抽出
    bat_players_name   = df_bat_concat['名前'].unique()
    pitch_players_name = df_pitch_concat['名前'].unique()
    
    #個人ごとにファイルを作成
    for bat_player_name in bat_players_name:
        df_bat = pd.DataFrame()
        df_bat = df_bat_concat[df_bat_concat['名前']==bat_player_name]
        # 必要なカラムのみ抽出
        df_bat = df_bat[['日付','試合形式','チーム', '打席', '打数', '安打', '単打',
                         '二塁打', '三塁打', '本塁打', '塁打','打点', '得点', '四球',
                         '死球', '犠打', '犠飛', '打撃妨害', '失策', '野選', '振り逃げ',
                         '三振', '併殺', '盗塁企画', '盗塁']]
        
        # 日付，試合形式でソート
        df_bat["日付"] = pd.to_datetime(df_bat['日付'])
        df_bat = df_bat.sort_values(['日付','試合形式'])
        df_bat["日付"]=df_bat['日付'].dt.strftime('%Y-%m-%d')
        # 試合数を保持
        bat_games = len(df_bat)
        # 合計行を作成
        df_bat.loc['合計'] = float('nan')
        df_bat.loc['合計', '打席':'盗塁'] = df_bat.iloc[:bat_games,3:].sum()
        # indexを設定
        df_bat = df_bat.set_index(['日付','試合形式','チーム'])
        
        mymodule.calc_bat_record(df_bat)
        
        is_pitch = False
        
        if bat_player_name in pitch_players_name:
            is_pitch = True
            df_pitch = df_pitch_concat[df_pitch_concat['名前']==bat_player_name]
            # 日付，試合形式でソート
            df_pitch["日付"] = pd.to_datetime(df_pitch['日付'])
            df_pitch = df_pitch.sort_values(['日付','試合形式'])
            df_pitch["日付"]=df_pitch['日付'].dt.strftime('%Y-%m-%d')
            # 試合数を保持
            pitch_games = len(df_pitch)
            # 合計行を作成
            df_pitch.loc['合計'] = float('nan')
            df_pitch.loc['合計', '完封':'自責点'] = df_pitch.iloc[0:pitch_games,3:].sum()
            # indexを設定
            df_pitch = df_pitch.set_index(['日付','試合形式','チーム'])
            df_pitch.drop('名前', axis=1, inplace=True)

            mymodule.calc_pitch_record(df_pitch)
            
        with pd.ExcelWriter(export_folder_path+'/'+bat_player_name+'.xlsx') as writer:
            df_bat.to_excel(writer, sheet_name='打撃成績')
            if is_pitch:
                df_pitch.to_excel(writer, sheet_name='投手成績')
                
        # 書式設定
        wb= px.load_workbook(export_folder_path+'/'+bat_player_name+'.xlsx')
        # 打撃成績
        ws_bat = wb.worksheets[0]
        set_format(ws_bat, bat_games, beginning_bat_rate, BAT_DARK_COLOR, BAT_THIN_COLOR)
        # 投手成績
        if is_pitch:
            ws_pitch = wb.worksheets[1]
            set_format(ws_pitch, pitch_games, beginning_pitch_rate, PITCH_DARK_COLOR, PITCH_THIN_COLOR)
            
        # 打撃の推移グラフを作成
        wb.create_sheet('打率推移')
        ws_avg = wb['打率推移']
        
        make_avg_graph_sheet(ws_bat, ws_avg)
                
        #保存
        wb.save(export_folder_path+'/'+bat_player_name+'.xlsx')
                
    # 出力したフォルダを開く
    subprocess.Popen(["explorer", export_folder_path], shell=True)
    print("success!")
    
if __name__ == '__main__':
    main()