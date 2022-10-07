import os
import pandas as pd
import datetime
import subprocess
import openpyxl as px
import mymodule
import sys
from openpyxl.xml.constants import NAMESPACES
from openpyxl.styles import Alignment, PatternFill

    
# 試合結果の全データを打撃/投手別で統合
def concat_games(paths, df_bat, df_pitch):
    for path in paths:
        df_bat_read_excel   = pd.read_excel(path, sheet_name='打撃成績', index_col=0)
        df_pitch_read_excel = pd.read_excel(path, sheet_name='投手成績', index_col=0)
        df_bat   = pd.concat([df_bat, df_bat_read_excel])
        df_pitch = pd.concat([df_pitch, df_pitch_read_excel])
    return df_bat, df_pitch

# 列の幅を設定
def set_column_width(ws: px.Workbook.worksheets, beginning_rate: int):
    name_width = 14
    width = 5
    ws.column_dimensions['A'].width = name_width
    for col in ws.iter_cols(min_col=2):
        column_num = col[0].column
        column_str = col[0].column_letter
        if column_num == beginning_rate:
            width = 7
        ws.column_dimensions[column_str].width = width
        
# 一行目を縦書きに設定
def set_vertical_writing_row1(ws: px.Workbook.worksheets):
    for row in ws.iter_rows(min_col=2,max_row=1):
        for cell in row:
            cell.alignment = Alignment(vertical='center', textRotation=255)
    
#　セルの背景色を設定
def set_backgroud_color(ws: px.Workbook.worksheets, dark_color: str, thin_color: str):
    max_row = ws.max_row
    # name
    for col in ws.iter_cols(min_row=2, max_row=max_row, max_col=1):  
        for cell in col:
            cell.fill = PatternFill(patternType = 'solid', fgColor=thin_color)
    # title
    for row in ws.iter_rows(max_row=1):
        for cell in row:
            cell.fill = PatternFill(patternType = 'solid', fgColor=dark_color)
    for row in ws.iter_rows(min_row=max_row):
        for cell in row:
            cell.fill = PatternFill(patternType = 'solid', fgColor=dark_color)

def main():
    
    BAT_DARK_COLOR = 'A4C6FF'
    BAT_THIN_COLOR = 'D9E5FF'

    PITCH_DARK_COLOR = 'FFA3A3'
    PITCH_THIN_COLOR = 'FFD9D9'
    
    beginning_bat_rate = 25
    beginning_pitch_rate = 18
    
    print("start")
    
    # パスを設定
    dirname = mymodule.get_dirname()
    import_foloder_path = dirname + '\試合結果'
    export_folder_path  = dirname + '\チーム成績'

    # 試合結果の全データを統合
    df_bat_concat   = pd.DataFrame()
    df_pitch_concat = pd.DataFrame()
    game_file_paths = mymodule.get_xlsx_file_paths(folder_path=import_foloder_path)
    df_bat_concat, df_pitch_concat = concat_games(paths=game_file_paths, df_bat=df_bat_concat, df_pitch=df_pitch_concat)
    
    # カラムに試合or登板を追加
    df_bat_concat['試合'] = 1
    df_pitch_concat['登板'] = 1

    # 各項目を合計したデータフレームを作成
    df_bat_sum = df_bat_concat[['試合', '打席', '打数', '安打', '単打', '二塁打', '三塁打', '本塁打', '塁打',
                                '打点', '得点', '四球', '死球', '犠打', '犠飛', '打撃妨害', '失策', '野選',
                                '振り逃げ', '三振', '併殺', '盗塁企画', '盗塁']].groupby('名前').sum()

    df_pitch_sum = df_pitch_concat[['登板', '完封', '完投', '勝利', '敗戦', '引き分け', 'セーブ', '奪アウト数', '投球数', '打者数', '被安打',
                                    '与四球', '与死球', '奪三振', '失点', '自責点']].groupby('名前').sum()

    # 成績を生成する選手を抽出
    players_name = mymodule.get_players_name(dirname)
    df_bat_sum   = df_bat_sum.filter(items=players_name, axis=0)
    df_pitch_sum = df_pitch_sum.filter(items=players_name, axis=0)

    # チーム総合をデータ化
    games = len(game_file_paths)
    df_bat_sum.loc['チーム総合']   = df_bat_sum.iloc[0:, 1:].sum()
    df_pitch_sum.loc['チーム総合'] = df_pitch_sum.iloc[0:, 1:].sum()
    df_bat_sum.loc['チーム総合', '試合']   = games
    df_pitch_sum.loc['チーム総合', '登板'] = games

    # 指標計算
    mymodule.calc_bat_record(df_bat_sum)
    mymodule.calc_pitch_record(df_pitch_sum)
    
    # ファイルに出力する
    date = str(datetime.date.today())
    with pd.ExcelWriter(export_folder_path + '/通算_' + date + '.xlsx') as writer:
        df_bat_sum.to_excel(writer, sheet_name='打撃成績')
        df_pitch_sum.to_excel(writer, sheet_name='投手成績')
        
    # 書式設定
    wb = px.load_workbook(export_folder_path+'/通算_'+date+'.xlsx')
    ws_bat = wb.worksheets[0]
    ws_pitch = wb.worksheets[1]
    
    # 列と行を固定化
    ws_bat.freeze_panes = 'B2'
    ws_pitch.freeze_panes = 'B2'
    
    # 試合数
    ws_bat['A1'] = str(games)+ '試合'
    ws_pitch['A1'] = str(games)+ '試合'
    
    # 率を小数第三位までに設定
    mymodule.set_rate_format(beginning=beginning_bat_rate, ws=ws_bat)
    mymodule.set_rate_format(beginning=beginning_pitch_rate, ws=ws_pitch)
        
    # 一行目を縦書きに設定
    set_vertical_writing_row1(ws_bat)
    set_vertical_writing_row1(ws_pitch)       
        
    # 列の幅を設定
    set_column_width(ws=ws_bat, beginning_rate=beginning_bat_rate)
    set_column_width(ws=ws_pitch, beginning_rate=beginning_pitch_rate)
    
    # ['A1']は横書き，中央揃えに
    ws_bat['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
    ws_pitch['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
    
    #　セルの塗りつぶし　1列目→1行目→最終行目
    set_backgroud_color(ws=ws_bat, dark_color=BAT_DARK_COLOR, thin_color= BAT_THIN_COLOR)
    set_backgroud_color(ws=ws_pitch, dark_color=PITCH_DARK_COLOR, thin_color=PITCH_THIN_COLOR)
    
    #保存
    wb.save(export_folder_path+'/通算_'+date+'.xlsx')
    
    # 出力したフォルダを開く
    subprocess.Popen(["explorer", export_folder_path], shell=True)

    print("success!")

if __name__ == '__main__':
    main()
