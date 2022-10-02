import os
from pathlib import Path
import openpyxl as px
# from openpyxl.xml.constants import NAMESPACES
import pandas as pd
import glob
# from openpyxl.styles import Alignment, PatternFill
import datetime
import subprocess

# 成績を生成する選手を取得
def get_players_name(path: str):
    wb = px.load_workbook(path + "\選手登録.xlsx")
    ws = wb.worksheets[0]
    for col in ws.columns:
        players_name = []
        for cell in col:
            players_name.append(cell.value)

    return players_name

# 打撃成績を計算
def calc_bat_record(df: pd.DataFrame):
    df['盗塁成功率'] = df['盗塁'] / df['盗塁企画']
    df['打率'] = df['安打'] / df['打数']
    divisor = df['打数'] + df['四球'] + df['死球'] + df['犠飛']
    df['出塁率'] = (df['安打'] + df['四球'] + df['死球']) / divisor
    df['長打率'] = df['塁打'] / df['打数']
    df['OPS'] = df['出塁率'] + df['長打率']
    df['BB/K'] = df['四球'] / df['三振']
    wOBA_divided = 0.692*df['四球'] + 0.73*df['死球'] + 0.966*df['失策'] + 0.865 *df['単打'] + 1.334*df['二塁打'] + 1.725*df['三塁打'] + 2.065*df['本塁打']
    df['wOBA'] = wOBA_divided / divisor
    
    #欠損値を0で補完
    df.fillna(0, inplace=True)
    
# 投手成績を計算
def calc_pitch_record(df: pd.DataFrame):
    # 投球回（Innings pitched / IP）
    IP = df['奪アウト数']/3
    # 奪アウト数　→　投球回数
    q, mod = divmod(df['奪アウト数'], 3)
    df['奪アウト数'] = q + 0.1*mod
    df.rename(columns={'奪アウト数':'投球回'}, inplace=True)
    df['防御率'] = df['自責点']*7 / IP
    df['奪三振率'] = df['奪三振']*7 / IP
    df['K%'] = df['奪三振'] / df['打者数']
    df['BB%'] = df['与四球'] / df['打者数']
    df['被打率'] = df['被安打'] / df['打者数']
    df['WHIP'] = (df['与四球'] + df['被安打']) / IP
    df['投球数/回'] = df['投球数']/IP
    
    #欠損値を0で補完
    df.fillna(0, inplace=True)


def main():
    print("開始")
    # パスを設定
    dirname = os.path.dirname(__file__)
    import_foloder_path = dirname + '\試合結果'
    export_folder_path  = dirname + '\チーム成績'

    # 試合結果のデータをdf_concatでつなぎ合わせる
    path = import_foloder_path + '\*.xlsx'
    file_paths = glob.glob(path)
    df_bat_concat   = pd.DataFrame()
    df_pitch_concat = pd.DataFrame()
    for path in file_paths:
        df_bat_read_excel   = pd.read_excel(path, sheet_name='打撃成績', index_col=0)
        df_pitch_read_excel = pd.read_excel(path, sheet_name='投手成績', index_col=0)
        df_bat_concat   = pd.concat([df_bat_concat, df_bat_read_excel])
        df_pitch_concat = pd.concat([df_pitch_concat, df_pitch_read_excel])

    # カラムに試合or登板を追加
    df_bat_concat['試合'] = 1
    df_pitch_concat['登板'] = 1

    # 集計データを合計したデータフレームを作成
    df_bat_sum = df_bat_concat[['試合', '打席', '打数', '安打', '単打', '二塁打', '三塁打', '本塁打', '塁打',
                                '打点', '得点', '四球', '死球', '犠打', '犠飛', '打撃妨害', '失策', '野選',
                                '振り逃げ', '三振', '併殺', '盗塁企画', '盗塁']].groupby('名前').sum()

    df_pitch_sum = df_pitch_concat[['登板', '完封', '完投', '勝利', '敗戦', '引き分け', 'セーブ', '奪アウト数', '投球数', '打者数', '被安打',
                                    '与四球', '与死球', '奪三振', '失点', '自責点']].groupby('名前').sum()

    # 成績を生成する選手を抽出
    players_name = get_players_name(dirname)
    df_bat_sum   = df_bat_sum.filter(items=players_name, axis=0)
    df_pitch_sum = df_pitch_sum.filter(items=players_name, axis=0)

    # チーム総合をデータ化
    games = len(file_paths)

    df_bat_sum.loc['チーム総合']   = df_bat_sum.iloc[0:, 1:].sum()
    df_pitch_sum.loc['チーム総合'] = df_pitch_sum.iloc[0:, 1:].sum()
    df_bat_sum.loc['チーム総合', '試合']   = games
    df_pitch_sum.loc['チーム総合', '登板'] = games

    # 指標計算
    calc_bat_record(df_bat_sum)
    calc_pitch_record(df_pitch_sum)
    
    #一旦ファイルに出力する
    date = str(datetime.date.today())
    with pd.ExcelWriter(export_folder_path + '/通算_' + date + '.xlsx') as writer:
        df_bat_sum.to_excel(writer, sheet_name='打撃成績')
        df_pitch_sum.to_excel(writer, sheet_name='投手成績')
        
    
    # 出力したフォルダを開く
    subprocess.Popen(["explorer", export_folder_path], shell=True)

    print("完了")

if __name__ == '__main__':
    main()


# # 書式設定

# wb = px.load_workbook(export_file_team_path+'/通算_'+date+'.xlsx')
# ws_bat = wb.worksheets[0]
# ws_pitch = wb.worksheets[1]

# bat_maxrow = ws_bat.max_row
# bat_maxcol = ws_bat.max_column

# pitch_maxrow = ws_pitch.max_row
# pitch_maxcol = ws_pitch.max_column


# #列名（アルファベット）を格納するリスト，excel_numberを作成
# def toAlpha2(num):
#     i = int((num-1)/26)
#     j = int(num-(i*26))
#     Alpha = ''
#     for z in i,j:
#         if z != 0:
#             Alpha += chr(z+64)
#     return Alpha

# bat_excel_number = []
# excel_number = []

# for i in range(1,bat_maxcol+1):
#     bat_excel_number.append(toAlpha2(i))

# for i in range(1,pitch_maxcol+1):
#     excel_number.append(toAlpha2(i))

# #列と行を固定化
# ws_bat.freeze_panes = 'B2'
# ws_pitch.freeze_panes = 'B2'

# # 列の幅を変更
# for i in range(0, bat_maxcol):
#     if i == 0:
#         width = 14
#     elif (1 <= i) & (i <= 24):
#         width = 4
#     else:
#         width = 8
#     ws_bat.column_dimensions[bat_excel_number[i]].width = width

# for i in range(0, pitch_maxcol):
#     if i == 0:
#         width = 14
#     elif (1 <= i) & (i <= 16):
#         width = 4
#     else:
#         width = 8
#     ws_pitch.column_dimensions[excel_number[i]].width = width

# #率を小数第三位までに設定
# #打撃
# for i in range(24,bat_maxcol):
#     for j in range(1,bat_maxrow+1):
#         ws_bat[bat_excel_number[i]+str(j)].number_format = "0.000"

# # 投手
# rate_format = "0.000"
# for i in range(17,pitch_maxcol):
#     if i == pitch_maxcol-1:
#         rate_format = "0.0"
#     for j in range(1,pitch_maxrow+1):
#         ws_pitch[excel_number[i]+str(j)].number_format = rate_format

# #データをセンタリング
# for i in range(1,bat_maxcol):
#     for j in range(2,bat_maxrow+1):
#         ws_bat[bat_excel_number[i]+str(j)].alignment = Alignment(horizontal = 'center')

# for i in range(1,pitch_maxcol):
#     for j in range(2,pitch_maxrow+1):
#         ws_pitch[excel_number[i]+str(j)].alignment = Alignment(horizontal = 'center')

# #ヘッダーを縦書きに設定
# for i in range(1,bat_maxcol):
#     ws_bat[bat_excel_number[i]+'1'].alignment = Alignment(vertical = 'center',textRotation = 255)

# for i in range(1,pitch_maxcol):
#     ws_pitch[excel_number[i]+'1'].alignment = Alignment(vertical = 'center',textRotation = 255)

# #試合数
# ws_bat['A1'] = str(games)+ '試合'
# ws_pitch['A1'] = str(games)+ '試合'

# # ['A1']は横書き，中央揃えに
# ws_bat['A1'].alignment = Alignment(horizontal = 'center',vertical='center')
# ws_pitch['A1'].alignment = Alignment(horizontal = 'center',vertical='center')

# #セルの塗りつぶし
# BAT_TITL_CELL_COLOR = 'A4C6FF'
# BAT_NAME_CELL_COLOR = 'D9E5FF'

# PITCH_TITL_CELL_COLOR = 'ffa3a3'
# PITCH_NAME_CELL_COLOR = 'ffd9d9'

# for i in range(0,bat_maxcol):
#     cell = ws_bat[bat_excel_number[i]+'1']
#     cell.fill = PatternFill(patternType = 'solid', fgColor=BAT_TITL_CELL_COLOR)

# for i in range(2,bat_maxrow):
#     cell = ws_bat['A'+ str(i)]
#     cell.fill = PatternFill(patternType = 'solid', fgColor=BAT_NAME_CELL_COLOR)

# for i in range(0,bat_maxcol):
#     cell = ws_bat[bat_excel_number[i]+str(bat_maxrow)]
#     cell.fill = PatternFill(patternType = 'solid', fgColor=BAT_TITL_CELL_COLOR)


# for i in range(0,pitch_maxcol):
#     cell = ws_pitch[excel_number[i]+'1']
#     cell.fill = PatternFill(patternType = 'solid', fgColor=PITCH_TITL_CELL_COLOR)

# for i in range(2,pitch_maxrow):
#     cell = ws_pitch['A'+ str(i)]
#     cell.fill = PatternFill(patternType = 'solid', fgColor=PITCH_NAME_CELL_COLOR)

# for i in range(0,pitch_maxcol):
#     cell = ws_pitch[excel_number[i]+str(pitch_maxrow)]
#     cell.fill = PatternFill(patternType = 'solid', fgColor=PITCH_TITL_CELL_COLOR)


# #保存
# wb.save(export_file_team_path+'/通算_'+date+'.xlsx')
